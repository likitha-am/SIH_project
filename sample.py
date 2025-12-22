# import pandas as pd
# import numpy as np
# import json, re
# from openpyxl import Workbook   # FIXED: correct import

# # ---------------------------------------------------
# # Load original dataset
# # ---------------------------------------------------
# IN_PATH = r"C:\Users\GIRIDHAR\OneDrive\Desktop\FRA\FRA_new_dataset.xlsx"
# OUT_PATH = r"C:\Users\GIRIDHAR\OneDrive\Desktop\FRA\FRA_new_dataset_augmented.xlsx"

# df = pd.read_excel(IN_PATH)

# # --- Robust parser ---
# def parse(cell):
#     if isinstance(cell, (list, tuple, np.ndarray)):
#         return np.array(cell, float)
#     s = str(cell)
#     try:
#         return np.array(json.loads(s), float)
#     except:
#         nums = re.findall(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", s)
#         return np.array([float(x) for x in nums], float)

# # Length of FRA vectors
# N = parse(df.loc[0, "magnitude_db"]).size
# rng = np.random.default_rng(42)

# # Synthetic fault classes
# new_faults = [
#     "winding_displacement_axial",
#     "winding_displacement_radial",
#     "inter_turn_short",
#     "open_circuit_winding",
#     "bushing_fault",
#     "tap_changer_fault"
# ]
# per_fault = 500  # samples per fault type

# def to_str(a):
#     return json.dumps(a.tolist())

# def interp_to_N(freqs, arr, N):
#     """Interpolate FRA array to target size N."""
#     if freqs.size < 2:
#         return np.full(N, arr[0] if arr.size else 0)
#     new_f = np.logspace(np.log10(freqs.min()), np.log10(freqs.max()), N)
#     order = np.argsort(freqs)
#     return np.interp(new_f, freqs[order], arr[order])

# # --- FRA synthesis ---
# def synthesize(freqs, mags, phs, ftype):
#     new_f = np.logspace(np.log10(freqs.min()), np.log10(freqs.max()), N)
#     logf = np.log10(new_f)

#     base = mags if mags.size == N else interp_to_N(freqs, mags, N)
#     phase = phs if phs.size == N else interp_to_N(freqs, phs, N)

#     if ftype == "winding_displacement_axial":
#         mu = np.median(logf) - 0.5
#         sigma = 0.7
#         a = 0.25
#         env = 1 + a * np.exp(-((logf - mu) ** 2) / (2 * sigma ** 2))
#         mags_new = base * env
#         phase += 5 * env

#     elif ftype == "winding_displacement_radial":
#         mu = np.median(logf) + 0.6
#         sigma = 0.6
#         a = -0.2
#         env = 1 + a * np.exp(-((logf - mu) ** 2) / (2 * sigma ** 2))
#         mags_new = base * env
#         phase += -5 * env

#     elif ftype == "inter_turn_short":
#         noise = rng.normal(0, 0.5, N)
#         dips = np.zeros(N)
#         for _ in range(3):
#             c = rng.uniform(logf.mean(), logf.max())
#             w = 0.1
#             d = 2
#             dips += d * np.exp(-((logf - c) ** 2) / (2 * w ** 2))
#         mags_new = base - dips + noise

#     elif ftype == "open_circuit_winding":
#         c = rng.uniform(logf.min(), logf.max())
#         w = 0.3
#         d = 5
#         mask = np.exp(-((logf - c) ** 2) / (2 * w ** 2))
#         mags_new = base - d * mask
#         phase += 20 * mask

#     elif ftype == "bushing_fault":
#         c = rng.uniform(logf.mean(), logf.max())
#         w = 0.03
#         d = 5
#         dip = d * np.exp(-((logf - c) ** 2) / (2 * w ** 2))
#         mags_new = base - dip

#     elif ftype == "tap_changer_fault":
#         ripple = 0.5 * np.sin(np.linspace(0, 6 * np.pi, N))
#         mags_new = base + ripple
#         phase += 2 * ripple

#     else:
#         mags_new = base.copy()

#     # Add measurement noise
#     mags_new += rng.normal(0, 0.05, N)
#     phase += rng.normal(0, 0.2, N)

#     return new_f, mags_new, phase

# # ---------------------------------------------------
# # Stream-write XLSX (write_only mode)
# # ---------------------------------------------------
# wb = Workbook(write_only=True)   # write-only mode = low RAM usage
# ws = wb.create_sheet("Sheet1")

# # Write header
# ws.append(df.columns.tolist())

# # Write original rows
# for _, r in df.iterrows():
#     ws.append(r.tolist())

# # Generate synthetic rows in chunks
# base_idx = rng.choice(df.shape[0], size=len(new_faults) * per_fault, replace=True)
# k = 0

# for ftype in new_faults:
#     for i in range(per_fault):

#         r = df.iloc[base_idx[k]]
#         k += 1

#         freqs = parse(r["frequency_hz"])
#         mags = parse(r["magnitude_db"])
#         phs = parse(r["phase_deg"])

#         f2, m2, p2 = synthesize(freqs, mags, phs, ftype)

#         new_row = r.copy()
#         new_row["frequency_hz"] = to_str(f2)
#         new_row["magnitude_db"] = to_str(m2)
#         new_row["phase_deg"] = to_str(p2)
#         new_row["fault_type"] = ftype
#         new_row["transformer_id"] = f"{r['transformer_id']}_synth_{ftype}_{i}"

#         ws.append(new_row.tolist())

# # Save final XLSX
# wb.save(OUT_PATH)

# print("\nDONE! Augmented XLSX saved to:\n", OUT_PATH)
import pandas as pd
import numpy as np
import json, re
from openpyxl import Workbook

# ---------------------------------------------------
# Load original dataset
# ---------------------------------------------------
IN_PATH = r"C:\Users\GIRIDHAR\OneDrive\Desktop\FRA\FRA_new_dataset.xlsx"
OUT_PATH = r"C:\Users\GIRIDHAR\OneDrive\Desktop\FRA\FRA_new_dataset_augmented_fixed.xlsx"

df = pd.read_excel(IN_PATH)

# --- parser ---
def parse(cell):
    if isinstance(cell, (list, tuple, np.ndarray)):
        return np.array(cell, float)
    s = str(cell)
    try:
        return np.array(json.loads(s), float)
    except:
        nums = re.findall(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", s)
        return np.array([float(x) for x in nums], float)

# FIX: Force global N to be the MAX length from original data
N = max(parse(df.loc[i, "magnitude_db"]).size for i in range(len(df)))
print("Using fixed FRA length:", N)

rng = np.random.default_rng(42)

new_faults = [
    "winding_displacement_axial",
    "winding_displacement_radial",
    "inter_turn_short",
    "open_circuit_winding",
    "bushing_fault",
    "tap_changer_fault"
]
per_fault = 500

def to_str(a):
    return json.dumps(a.tolist())

def interp_to_N(freqs, arr):
    """Interpolation to fixed N length"""
    if freqs.size < 2:
        return np.full(N, arr[0] if arr.size else 0)
    new_f = np.logspace(np.log10(freqs.min()), np.log10(freqs.max()), N)
    order = np.argsort(freqs)
    return new_f, np.interp(new_f, freqs[order], arr[order])

def synthesize(freqs, mags, phs, ftype):
    # Always interpolate to length N
    new_f, base = interp_to_N(freqs, mags)
    _, phase = interp_to_N(freqs, phs)

    logf = np.log10(new_f)

    if ftype == "winding_displacement_axial":
        mu = np.median(logf) - 0.5
        env = 1 + 0.25 * np.exp(-((logf - mu) ** 2) / (2 * 0.7 ** 2))
        mags_new = base * env
        phase += 5 * env

    elif ftype == "winding_displacement_radial":
        mu = np.median(logf) + 0.6
        env = 1 - 0.2 * np.exp(-((logf - mu) ** 2) / (2 * 0.6 ** 2))
        mags_new = base * env
        phase -= 5 * env

    elif ftype == "inter_turn_short":
        dips = np.zeros(N)
        for _ in range(3):
            c = rng.uniform(logf.mean(), logf.max())
            dips += 2 * np.exp(-((logf - c) ** 2) / (2 * 0.1 ** 2))
        mags_new = base - dips + rng.normal(0, 0.5, N)

    elif ftype == "open_circuit_winding":
        c = rng.uniform(logf.min(), logf.max())
        mask = np.exp(-((logf - c) ** 2) / (2 * 0.3 ** 2))
        mags_new = base - 5 * mask
        phase += 20 * mask

    elif ftype == "bushing_fault":
        c = rng.uniform(logf.mean(), logf.max())
        dip = 5 * np.exp(-((logf - c) ** 2) / (2 * 0.03 ** 2))
        mags_new = base - dip

    elif ftype == "tap_changer_fault":
        ripple = 0.5 * np.sin(np.linspace(0, 6 * np.pi, N))
        mags_new = base + ripple
        phase += 2 * ripple

    else:
        mags_new = base.copy()

    mags_new += rng.normal(0, 0.05, N)
    phase += rng.normal(0, 0.2, N)

    return new_f, mags_new, phase

# ---------------------------------------------------
# Write streaming XLSX
# ---------------------------------------------------
wb = Workbook(write_only=True)
ws = wb.create_sheet("Sheet1")

ws.append(df.columns.tolist())

# Original rows
for _, r in df.iterrows():
    ws.append(r.tolist())

# Synthetic rows
base_idx = rng.choice(df.shape[0], size=len(new_faults) * per_fault, replace=True)

k = 0
for ftype in new_faults:
    for i in range(per_fault):
        r = df.iloc[base_idx[k]]
        k += 1

        freqs = parse(r["frequency_hz"])
        mags = parse(r["magnitude_db"])
        phs = parse(r["phase_deg"])

        f2, m2, p2 = synthesize(freqs, mags, phs, ftype)

        new_row = r.copy()
        new_row["frequency_hz"] = to_str(f2)
        new_row["magnitude_db"] = to_str(m2)
        new_row["phase_deg"] = to_str(p2)
        new_row["fault_type"] = ftype

        new_row["transformer_id"] = f"{r['transformer_id']}_synth_{ftype}_{i}"

        ws.append(new_row.tolist())

wb.save(OUT_PATH)

print("\nDONE — saved fixed file to:\n", OUT_PATH)
